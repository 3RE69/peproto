from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import load_workbook
import os
import tempfile

app = Flask(__name__)

@app.route("/")
def home():
    return render_template("peproto.html")  # Serves the HTML page

@app.route("/process", methods=["POST"])
def process_files():
    try:
        source_file = request.files.get("sourceFile")  # Get Source File
        reference_file = request.files.get("referenceFile")  # Get Reference File
        format_type = request.form.get("formatType")  # Identify format type

        if not source_file or not reference_file or not format_type:
            return jsonify({"status": "error", "message": "Please upload both Source and Reference files."}), 400

        temp_dir = tempfile.mkdtemp()
        source_path = os.path.join(temp_dir, source_file.filename)
        reference_path = os.path.join(temp_dir, reference_file.filename)

        source_file.save(source_path)
        reference_file.save(reference_path)

        output_path = os.path.join(temp_dir, "processed_output.xlsx")

        if format_type.lower() == "production":
            message = process_production_format(source_path, reference_path, output_path)
        elif format_type.lower() == "tool":
            message = process_tool_format(source_path, reference_path, output_path)
        else:
            return jsonify({"status": "error", "message": "Invalid format type."}), 400

        return jsonify({"status": "success", "message": message, "downloadUrl": f"/download?path={output_path}"})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/download")
def download_file():
    file_path = request.args.get("path")
    if file_path and os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    return jsonify({"status": "error", "message": "File not found"}), 404

def process_production_format(source_path, reference_path, output_path):
    """Processes Production Format using the uploaded reference file"""
    try:
        wb1 = load_workbook(source_path)
        sheet1 = wb1.active

        wb2 = load_workbook(reference_path)  # Use uploaded reference file
        sheet2 = wb2.active

        combined_data = [" ".join([str(sheet1.cell(row=i, column=j).value) for j in range(1, 6)])
                         for i in range(2, sheet1.max_row + 1)]

        for merged_range in list(sheet2.merged_cells.ranges):
            if merged_range.min_col == 2:
                sheet2.unmerge_cells(str(merged_range))

        row_number = 27
        for value in combined_data:
            sheet2.cell(row=row_number, column=2).value = value
            row_number += 12

        wb2.save(output_path)
        return "Production format processing complete. Click download."

    except Exception as e:
        return f"Error during processing: {str(e)}"

def process_tool_format(source_path, reference_path, output_path):
    """Processes Tool Format using the uploaded reference file"""
    try:
        wb1 = load_workbook(source_path)
        sheet1 = wb1.active

        wb2 = load_workbook(reference_path)
        sheet2 = wb2.active

        # Identify column indexes in source file based on headers
        header_row = 1  # Assuming headers are in row 1
        source_headers = {cell.value.strip(): idx + 1 for idx, cell in enumerate(sheet1[header_row]) if cell.value}

        # Define columns to extract
        columns_needed = ["UID", "PROJECT", "PART NO", "DESCRIPTION"]
        source_indices = [source_headers[col] for col in columns_needed if col in source_headers]

        # Extract data from source file (starting from row 2)
        data_to_write = [[sheet1.cell(row=i, column=col_idx).value for col_idx in source_indices] 
                         for i in range(2, sheet1.max_row + 1)]

        # Unmerge any merged cells in the target sheet
        for merged_range in list(sheet2.merged_cells.ranges):
            sheet2.unmerge_cells(str(merged_range))

        # Define target columns (A, B, D, F)
        target_cols = [1, 2, 4, 6]  # Corresponding column numbers in sheet2

        # Write extracted data sequentially from row 4 onwards
        row_number = 4  
        for row_values in data_to_write:
            for col_idx, value in zip(target_cols, row_values):
                sheet2.cell(row=row_number, column=col_idx).value = value
            row_number += 1  # Move to the next row

        wb2.save(output_path)
        return "Tool format processing complete. Click download."

    except Exception as e:
        return f"Error during processing: {str(e)}"

if __name__ == "__main__":
    app.run(debug=True)